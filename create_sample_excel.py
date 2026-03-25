"""
Genera un Excel de muestra con formato EERR Ascent.
Ejecutar: python create_sample_excel.py
Produce: eerr_sample.xlsx
"""

from __future__ import annotations

import io
import random

import pandas as pd


# ----------------------------------------------------------------
# Definición del EERR de muestra (Distribuidora Ejemplo SA)
# ----------------------------------------------------------------
LINES = [
    # (código, nombre, tag, es_subtotal)
    (1000, "Ventas Netas",                    "Ingresos",   True),
    (1100, "Ventas Producto A",               "Ingresos",   False),
    (1200, "Ventas Producto B",               "Ingresos",   False),
    (1300, "Ventas Producto C",               "Ingresos",   False),
    (2000, "Costo de Mercadería Vendida",     "Costos",     True),
    (2100, "Costo Producto A",                "Costos",     False),
    (2200, "Costo Producto B",                "Costos",     False),
    (2300, "Costo Producto C",                "Costos",     False),
    (3000, "Utilidad Bruta",                  "KPI",        True),
    (4000, "Gastos Operativos",               "Gastos",     True),
    (4100, "Sueldos y Cargas Sociales",       "Gastos",     False),
    (4200, "Alquileres",                      "Gastos",     False),
    (4300, "Servicios y Utilities",           "Gastos",     False),
    (4400, "Publicidad y Marketing",          "Gastos",     False),
    (4500, "Logística y Distribución",        "Gastos",     False),
    (5000, "EBITDA",                          "KPI",        True),
    (5001, "Margen EBITDA %",                 "KPI",        False),
    (6000, "Depreciaciones y Amortizaciones", "Gastos",     False),
    (7000, "Resultado Operativo (EBIT)",      "KPI",        True),
    (8000, "Resultado Financiero",            "Financiero", True),
    (8100, "Intereses Bancarios",             "Financiero", False),
    (8200, "Diferencias de Cambio",           "Financiero", False),
    (9000, "Resultado antes de Impuestos",    "KPI",        True),
    (9100, "Impuesto a las Ganancias",        "Impuestos",  False),
    (10000, "Utilidad Neta",                  "KPI",        True),
    (10001, "Margen Neto %",                  "KPI",        False),
]

MONTHS = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
          "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

# Valores base mensuales 2025 para Ventas Producto A, B, C (ARS)
_BASE_A_ARS = [300_000, 310_000, 325_000, 340_000, 355_000, 370_000,
               385_000, 395_000, 410_000, 425_000, 440_000, 455_000]
_BASE_B_ARS = [200_000, 210_000, 218_000, 225_000, 233_000, 241_000,
               249_000, 256_000, 263_000, 271_000, 280_000, 289_000]
_BASE_C_ARS = [100_000, 105_000, 110_000, 115_000, 120_000, 125_000,
               130_000, 135_000, 140_000, 145_000, 150_000, 155_000]

FX_2025 = 1_000  # ARS por USD
FX_2024 = 850


def _add_noise(vals: list[float], pct: float = 0.03) -> list[float]:
    return [v * (1 + random.uniform(-pct, pct)) for v in vals]


def _build_monthly(year: int) -> dict[int, list[float]]:
    """Construye valores mensuales por código para el año dado."""
    factor = 1.0 if year == 2025 else 0.75   # 2024 ~25% menos

    a = [v * factor for v in _add_noise(_BASE_A_ARS)]
    b = [v * factor for v in _add_noise(_BASE_B_ARS)]
    c = [v * factor for v in _add_noise(_BASE_C_ARS)]
    ventas = [a[i] + b[i] + c[i] for i in range(12)]

    costo_pct = 0.40
    ca = [-v * costo_pct for v in a]
    cb = [-v * costo_pct for v in b]
    cc = [-v * costo_pct for v in c]
    cmv = [ca[i] + cb[i] + cc[i] for i in range(12)]

    ub = [ventas[i] + cmv[i] for i in range(12)]  # Utilidad Bruta

    sueldos = [-v * 0.16 for v in ventas]
    alq     = [-v * 0.06 for v in ventas]
    serv    = [-v * 0.04 for v in ventas]
    pub     = [-v * 0.04 for v in ventas]
    logist  = [-v * 0.04 for v in ventas]
    go      = [sueldos[i] + alq[i] + serv[i] + pub[i] + logist[i] for i in range(12)]

    ebitda  = [ub[i] + go[i] for i in range(12)]
    marg_ebitda = [ebitda[i] / ventas[i] * 100 if ventas[i] else 0 for i in range(12)]

    depre   = [-ventas[i] * 0.03 for i in range(12)]
    ebit    = [ebitda[i] + depre[i] for i in range(12)]

    interes = [-ventas[i] * 0.04 for i in range(12)]
    fx_diff = [-ventas[i] * 0.01 for i in range(12)]
    rfin    = [interes[i] + fx_diff[i] for i in range(12)]

    eai     = [ebit[i] + rfin[i] for i in range(12)]
    imp     = [eai[i] * -0.35 if eai[i] > 0 else 0 for i in range(12)]
    un      = [eai[i] + imp[i] for i in range(12)]
    marg_n  = [un[i] / ventas[i] * 100 if ventas[i] else 0 for i in range(12)]

    go_total = go

    return {
        1000: ventas,
        1100: a,
        1200: b,
        1300: c,
        2000: cmv,
        2100: ca,
        2200: cb,
        2300: cc,
        3000: ub,
        4000: go_total,
        4100: sueldos,
        4200: alq,
        4300: serv,
        4400: pub,
        4500: logist,
        5000: ebitda,
        5001: marg_ebitda,
        6000: depre,
        7000: ebit,
        8000: rfin,
        8100: interes,
        8200: fx_diff,
        9000: eai,
        9100: imp,
        10000: un,
        10001: marg_n,
    }


def _build_df(year: int, currency: str) -> pd.DataFrame:
    monthly = _build_monthly(year)
    fx = FX_2025 if year == 2025 else FX_2024

    rows = []
    for code, name, tag, _ in LINES:
        vals_ars = monthly[code]
        if currency == "USD":
            vals = [v / fx for v in vals_ars]
        else:
            vals = vals_ars

        q1 = sum(vals[0:3])
        q2 = sum(vals[3:6])
        q3 = sum(vals[6:9])
        q4 = sum(vals[9:12])
        anio = sum(vals)

        row: dict = {"Código": code, "Nombre": name, "Tag": tag}
        for i, m in enumerate(MONTHS):
            row[m] = round(vals[i], 2)
        row["1 Trim"] = round(q1, 2)
        row["2 Trim"] = round(q2, 2)
        row["3 Trim"] = round(q3, 2)
        row["4 Trim"] = round(q4, 2)
        row["Año"] = round(anio, 2)
        rows.append(row)

    return pd.DataFrame(rows)


def create_sample_excel() -> bytes:
    """Genera el Excel en memoria y retorna bytes."""
    random.seed(42)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for year in (2025, 2024):
            for currency in ("ARS", "USD"):
                sheet_name = f"EERR {year} {currency}"
                df = _build_df(year, currency)
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                # Estilos básicos con openpyxl
                ws = writer.sheets[sheet_name]
                from openpyxl.styles import Alignment, Font, PatternFill

                header_fill = PatternFill("solid", fgColor="2D2D5E")
                subtotal_fill = PatternFill("solid", fgColor="E8E8F5")
                header_font = Font(color="FFFFFF", bold=True)

                # Encabezado
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

                # Subtotales
                subtotal_codes = {c for c, *_ in LINES if _[1]}  # is_subtotal
                for row_i, (code, *_) in enumerate(LINES, start=2):
                    is_sub = _[1]
                    if is_sub:
                        for cell in ws[row_i]:
                            cell.fill = subtotal_fill
                            cell.font = Font(bold=True)

                # Ancho de columnas
                ws.column_dimensions["A"].width = 10
                ws.column_dimensions["B"].width = 38
                ws.column_dimensions["C"].width = 14
                for col in ws.iter_cols(min_col=4, max_col=ws.max_column):
                    ws.column_dimensions[col[0].column_letter].width = 14

    return buf.getvalue()


if __name__ == "__main__":
    data = create_sample_excel()
    with open("eerr_sample.xlsx", "wb") as f:
        f.write(data)
    print("✅ eerr_sample.xlsx generado correctamente.")
    print("   Hojas: EERR 2025 ARS | EERR 2024 ARS | EERR 2025 USD | EERR 2024 USD")
